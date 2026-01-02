import hashlib
import json
import logging
import math
import os
import re
import sys
from pathlib import Path
from typing import Dict, List, Tuple

from bs4 import BeautifulSoup
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from seleniumbase import SB

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")


class UdemyScraper:
    """Scraper for Udemy courses to extract course information."""

    # Constants
    COURSES_PER_PAGE = 12
    DEFAULT_TIMEOUT = 15
    RETRY_WAIT_SECONDS = 60
    MAX_LOGIN_ATTEMPTS = 3
    PASSWORDLESS_WAIT_SECONDS = 60
    SAVE_FREQUENCY = 10  # Save Excel after every N courses

    def __init__(self, email: str, password: str, account_name: str):
        """Initialize the scraper with user credentials.

        Args:
            email: User's email address
            password: User's password
            account_name: Account first name for verification
        """
        self.email = email
        self.password = password
        self.account_name = account_name
        self.logger = logging.getLogger(__name__)

        # Create saved_cookies directory if it doesn't exist
        self.cookies_dir = Path("saved_cookies")
        self.cookies_dir.mkdir(exist_ok=True)
        self.cookies_file = self.cookies_dir / "cookies.txt"

        self.cache_dir = Path("cache_files")
        self.cache_dir.mkdir(exist_ok=True)
        self.course_urls_cache = self.cache_dir / "course_urls_cache.json"
        self.ignored_courses_file = self.cache_dir / "ignored_courses.txt"
        self.courses_xlsx_file = Path("udemy_courses.xlsx")

    def login(self, force: bool = False) -> bool:
        """Authenticate with Udemy.

        Args:
            force: If True, ignore existing cookies and force new login

        Returns:
            bool: True if login was successful, False otherwise
        """
        with SB(uc=True, test=True, locale_code="en") as sb:
            login_url = "https://www.udemy.com/join/passwordless-auth/?locale=en_US&next=https%3A%2F%2Fwww.udemy.com%2F&response_type=html&action=login&mode"
            login_selector = "#form-group--1"
            passwordless_button_selector_1 = (
                "button[type='submit']:contains('Continue')"
            )
            passwordless_button_selector_2 = "button[type='submit']:contains('Log in')"
            password_selector = "#form-group--3"
            passwordbased_button_selector = "#udemy > div.ud-main-content-wrapper > div.ud-main-content > div > main > div > div > div:nth-child(2) > form > button"

            try:
                sb.uc_open_with_reconnect(login_url, 5)  # open url bypassing captcha
                if (
                    not force and self.cookies_file.exists()
                ):  # use existing cookies if available
                    sb.load_cookies(str(self.cookies_file))
                    self.logger.info("LOADED COOKIES")
                    sb.refresh()
                    sb.assert_text_visible(self.account_name)
                    return True
                else:
                    raise ValueError("FORCE LOGIN OR NO COOKIES AVAILABLE")
            except (ValueError, AssertionError) as e:  # not logged in or forced login
                self.logger.debug(f"COOKIE LOGIN FAILED: {e}")
                return self._perform_login(
                    sb,
                    login_url,
                    login_selector,
                    passwordless_button_selector_1,
                    passwordless_button_selector_2,
                    password_selector,
                    passwordbased_button_selector,
                )

    def _perform_login(
        self,
        sb,
        login_url: str,
        login_selector: str,
        passwordless_button_selector_1: str,
        passwordless_button_selector_2: str,
        password_selector: str,
        passwordbased_button_selector: str,
    ) -> bool:
        """Handle the actual login process.

        Returns:
            bool: True if login successful, False otherwise
        """
        current_attempt = 0

        while current_attempt < self.MAX_LOGIN_ATTEMPTS:
            current_attempt += 1
            try:
                sb.uc_open_with_reconnect(login_url, 5)
                sb.get_element(login_selector, timeout=self.DEFAULT_TIMEOUT).click()
                sb.type(login_selector, self.email)

                try:  # try passwordless login first
                    sb.get_element(
                        passwordless_button_selector_1, timeout=self.DEFAULT_TIMEOUT
                    ).click()
                    self.logger.warning(
                        f"ENTER LOGIN CODE FROM EMAIL AND WAIT {self.PASSWORDLESS_WAIT_SECONDS}s FOR AUTO CLICK!"
                    )
                    sb.wait(self.PASSWORDLESS_WAIT_SECONDS)
                    sb.get_element(
                        passwordless_button_selector_2, timeout=self.DEFAULT_TIMEOUT
                    ).click()
                except Exception as e:  # fall back to password-based login
                    self.logger.debug(f"PASSWORDLESS LOGIN UNAVAILABLE: {e}")
                    sb.get_element(
                        password_selector, timeout=self.DEFAULT_TIMEOUT
                    ).click()
                    sb.type(password_selector, self.password)
                    sb.get_element(
                        passwordbased_button_selector, timeout=self.DEFAULT_TIMEOUT
                    ).click()

                # Verify login success
                sb.assert_text_visible(self.account_name)
                sb.save_cookies(str(self.cookies_file))
                return True

            except Exception as e:
                self.logger.warning(
                    f"LOGIN FAILED (ATTEMPT {current_attempt}/{self.MAX_LOGIN_ATTEMPTS}): {e}"
                )
                if current_attempt < self.MAX_LOGIN_ATTEMPTS:
                    self.logger.info(
                        f"WAITING {self.RETRY_WAIT_SECONDS}s BEFORE RETRYING..."
                    )
                    sb.wait(self.RETRY_WAIT_SECONDS)
                else:
                    self.logger.error(
                        f"FAILED TO LOGIN AFTER {self.MAX_LOGIN_ATTEMPTS} ATTEMPTS"
                    )
                    return False

    def scrape_courses(self) -> Dict[str, str]:
        """Scrape all user's enrolled courses.

        Returns:
            Dict mapping course titles to course duration
        """
        with SB(uc=True, test=True, locale_code="en") as sb:
            courses_url = "https://www.udemy.com/home/my-courses/learning/"

            # Selectors
            overview_selector = "//span[text()='Overview']"
            time_selector = "//div[contains(text(), 'Video:')]"
            pagination_selector = "div.pagination--pagination-label--tgma-"
            course_grid_selector = "div.my-courses__course-card-grid"

            sb.uc_open_with_reconnect(courses_url, 5)
            sb.load_cookies(str(self.cookies_file))
            sb.refresh()
            sb.get_element(pagination_selector, timeout=self.DEFAULT_TIMEOUT)

            # Get total number of courses and pages
            courses_count, pages_count = self._get_course_metadata(sb)
            self.logger.info(f"TOTAL COURSES: {courses_count}")

            # Get all course URLs with pagination
            courses_list = self._get_all_course_urls(
                sb, courses_url, course_grid_selector, pages_count
            )

            # Get existing courses to avoid re-scraping
            existing_courses = self._get_existing_courses()

            # Get ignored courses (non-video courses)
            ignored_courses = self._get_ignored_courses()

            # Scrape details for each course
            courses_details = self._scrape_course_details(
                sb,
                courses_list,
                existing_courses,
                ignored_courses,
                overview_selector,
                time_selector,
            )

            self.logger.info("SAVED COURSE DETAILS!")
            return courses_details

    def _get_course_metadata(self, sb) -> Tuple[int, int]:
        """Extract course count and pagination info.

        Returns:
            Tuple of (course count, page count)
        """
        soup = BeautifulSoup(sb.get_page_source(), "lxml")
        courses_num_details = soup.select("div[class*='pagination-label']")[
            0
        ].text.strip()
        courses_num_match = re.search(r"of (\d+) courses", courses_num_details)
        courses_count = int(courses_num_match.group(1))
        pages_count = math.ceil(courses_count / self.COURSES_PER_PAGE)
        return courses_count, pages_count

    def _get_all_course_urls(
        self, sb, courses_url: str, course_grid_selector: str, pages_count: int
    ) -> List[str]:
        """Get URLs for all courses from pagination.

        Returns:
            List of course URLs
        """
        if self.course_urls_cache.exists():
            self.logger.info("LOADING COURSE URLS FROM CACHE")
            with open(self.course_urls_cache, "r", encoding="utf-8") as f:
                cached_data = json.load(f)
                if cached_data.get("pages_count") == pages_count:
                    self.logger.info(
                        f"LOADED {len(cached_data['urls'])} COURSE URLS FROM CACHE"
                    )
                    return cached_data["urls"]
                else:
                    self.logger.info("CACHE INVALID (PAGE COUNT CHANGED), RE-SCRAPING")

        courses_list = []

        for page in range(1, pages_count + 1):
            sb.uc_open(f"{courses_url}?p={page}")
            sb.get_element(course_grid_selector, timeout=self.DEFAULT_TIMEOUT)

            soup = BeautifulSoup(sb.get_page_source(), "lxml")
            course_elements = soup.find_all(
                "h3", attrs={"data-purpose": "course-title-url"}
            )

            for element in course_elements:
                course_link = element.find("a").get("href")
                courses_list.append(f"https://www.udemy.com{course_link}")

            self.logger.info(f"PROCESSED PAGE #{page}/{pages_count}")

        cache_data = {"pages_count": pages_count, "urls": courses_list}
        with open(self.course_urls_cache, "w", encoding="utf-8") as f:
            json.dump(cache_data, f, indent=2)
        self.logger.info(f"CACHED {len(courses_list)} COURSE URLS")

        return courses_list

    def _get_existing_courses(self) -> dict:
        """Get set of courses that have already been scraped.

        Returns:
            Dict mapping course URLs to their row numbers in XLSX
        """
        existing_courses = {}

        if not self.courses_xlsx_file.exists():
            return existing_courses

        try:
            wb = load_workbook(self.courses_xlsx_file)
            ws = wb.active

            # Skip header row and read course URLs from first column with row numbers
            for idx, row in enumerate(
                ws.iter_rows(min_row=2, values_only=True), start=2
            ):
                if row and row[0]:  # Ensure row is not empty and has URL
                    existing_courses[row[0]] = idx

            wb.close()
            self.logger.info(
                f"LOADED {len(existing_courses)} EXISTING COURSES FROM XLSX"
            )
        except Exception as e:
            self.logger.error(f"ERROR READING XLSX FILE: {e}")

        return existing_courses

    def _get_ignored_courses(self) -> set:
        """Get set of courses that have been ignored (non-video courses).

        Returns:
            Set of course URLs that are ignored
        """
        ignored_courses = set()
        if not self.ignored_courses_file.exists():
            return ignored_courses

        with open(self.ignored_courses_file, "r", encoding="utf-8") as file:
            for line in file:
                line = line.strip()
                if line:  # Ensure line is not empty
                    ignored_courses.add(line)

        return ignored_courses

    def _get_cache_path(self, course_url: str) -> Path:
        """Get cache file path for a course URL.

        Args:
            course_url: The course URL to cache

        Returns:
            Path to the cache file
        """
        url_hash = hashlib.md5(course_url.encode()).hexdigest()
        return self.cache_dir / f"{url_hash}.html"

    def _scrape_course_details(
        self,
        sb,
        courses_list: List[str],
        existing_courses: set,
        ignored_courses: set,
        overview_selector: str,
        time_selector: str,
    ) -> Dict[str, str]:
        """Scrape details for each course.

        Returns:
            Dict mapping course titles to durations
        """
        courses_details = {}

        # Load or create Excel workbook
        if self.courses_xlsx_file.exists():
            wb = load_workbook(self.courses_xlsx_file)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(["Course Link", "Course Title", "Course Time"])

        processed_count = 0
        try:
            for index, course_url in enumerate(courses_list, 1):
                if course_url in existing_courses:
                    self.logger.info(f"SKIPPED COURSE #{index} (EXISTING)")
                    continue

                if course_url in ignored_courses:
                    self.logger.info(f"SKIPPED COURSE #{index} (IGNORED - NON-VIDEO)")
                    continue

                try:
                    cache_file = self._get_cache_path(course_url)

                    if cache_file.exists():
                        self.logger.info(f"LOADING COURSE #{index} FROM CACHE")
                        with open(cache_file, "r", encoding="utf-8") as f:
                            html_content = f.read()
                        soup = BeautifulSoup(html_content, "lxml")
                    else:
                        sb.uc_open(course_url)
                        sb.get_element(
                            overview_selector, timeout=self.DEFAULT_TIMEOUT
                        ).click()

                        # Try to find video element, if not found it's a non-video course
                        try:
                            sb.get_element(time_selector, timeout=2)
                            html_content = sb.get_page_source()
                            cache_file.write_text(html_content, encoding="utf-8")
                            self.logger.info(f"CACHED COURSE #{index} PAGE DATA")
                            soup = BeautifulSoup(html_content, "lxml")
                        except Exception:
                            # No video element found - mark as ignored
                            with open(
                                self.ignored_courses_file, "a", encoding="utf-8"
                            ) as ignore_file:
                                ignore_file.write(f"{course_url}\n")
                            self.logger.info(
                                f"IGNORED COURSE #{index} (NON-VIDEO COURSE)"
                            )
                            continue

                    # Extract course title
                    course_title_element = soup.find("title")
                    course_title = (
                        course_title_element.text.strip()
                        .replace("Course: ", "")
                        .replace(" | Udemy", "")
                        if course_title_element
                        else "N/A"
                    )

                    # Extract course duration
                    video_length_div = soup.find(
                        "div", class_=re.compile("video-length")
                    )
                    course_time_element = video_length_div.find(class_="ud-heading-md")
                    course_time = (
                        course_time_element.text.strip()
                        if course_time_element
                        else "N/A"
                    )

                    # Save to Excel and dict
                    ws.append([course_url, course_title, course_time])
                    courses_details[course_title] = course_time
                    processed_count += 1

                    # Save periodically for reliability
                    if processed_count % self.SAVE_FREQUENCY == 0:
                        wb.save(self.courses_xlsx_file)
                        self.logger.info(f"AUTO-SAVED AFTER {processed_count} COURSES")

                    self.logger.info(f"PROCESSED COURSE #{index} (NEW)")

                except Exception as e:
                    self.logger.error(f"ERROR PROCESSING COURSE {course_url}: {str(e)}")
        finally:
            # Always save the workbook
            wb.save(self.courses_xlsx_file)
            wb.close()
            self.logger.info(f"SAVED TO {self.courses_xlsx_file}")

        return courses_details

    def format_xlsx(self) -> None:
        """Format course durations from text to minutes."""
        if not self.courses_xlsx_file.exists():
            self.logger.error(f"INPUT FILE {self.courses_xlsx_file} NOT FOUND")
            return

        try:
            wb = load_workbook(self.courses_xlsx_file)
            ws = wb.active

            total_minutes = 0
            course_count = 0

            # Skip header row, iterate with row objects to modify cells
            for row in ws.iter_rows(min_row=2):
                if not row[2].value:  # Skip rows without time
                    continue

                time_str = str(row[2].value)
                minutes = self._convert_to_minutes(time_str)
                row[2].value = minutes  # Update the cell with integer minutes

                if minutes > 0:
                    total_minutes += minutes
                    course_count += 1

            # Save the updated workbook
            wb.save(self.courses_xlsx_file)
            wb.close()

            if course_count > 0:
                hours = total_minutes / 60
                self.logger.info(
                    f"TOTAL COURSE TIME: {hours:.1f} HOURS ({total_minutes} MINUTES)"
                )
                self.logger.info(
                    f"AVERAGE COURSE LENGTH: {(total_minutes / course_count):.1f} MINUTES"
                )

            self.logger.info("FORMATTED COURSE DETAILS!")

        except Exception as e:
            self.logger.error(f"ERROR FORMATTING XLSX: {e}")

    def _convert_to_minutes(self, time_str: str) -> int:
        """Convert time string to minutes.

        Args:
            time_str: String like "5 hours 30 mins" or "45 mins"

        Returns:
            Total minutes as integer
        """
        if time_str == "N/A":
            return 0

        total_minutes = 0
        hours_match = re.search(r"(\d+(?:\.\d+)?)\s*hours?", time_str)
        minutes_match = re.search(r"(\d+)\s*mins?", time_str)

        if hours_match:
            total_minutes += float(hours_match.group(1)) * 60
        if minutes_match:
            total_minutes += int(minutes_match.group(1))

        return int(total_minutes)


def main():
    """Main entry point for the script."""
    try:
        load_dotenv()  # Load environment variables from .env file

        email = os.getenv("UDEMY_EMAIL")
        password = os.getenv("UDEMY_PASSWORD")
        account_name = os.getenv("UDEMY_ACCOUNT_NAME")
        force_login = os.getenv("FORCE_LOGIN", "False").lower() == "true"

        if not all([email, password, account_name]):
            print("ERROR: Missing required environment variables:")
            print("  - UDEMY_EMAIL")
            print("  - UDEMY_PASSWORD")
            print("  - UDEMY_ACCOUNT_NAME")
            sys.exit(1)

        scraper = UdemyScraper(email, password, account_name)
        scraper.logger.info("=== Udemy Course Scraper ===")

        # Attempt login
        if scraper.login(force=force_login):
            scraper.logger.info("LOGGED IN SUCCESSFULLY!")
            scraper.scrape_courses()  # Scrape courses
            scraper.format_xlsx()  # Format XLSX for analysis
            scraper.logger.info("SCRAPING COMPLETED SUCCESSFULLY!")
        else:
            scraper.logger.error("LOGIN UNSUCCESSFUL!")
            sys.exit(1)

    except KeyboardInterrupt:
        scraper.logger.info("OPERATION CANCELLED BY USER")
        sys.exit(0)
    except Exception as e:
        scraper.logger.error(f"{e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
